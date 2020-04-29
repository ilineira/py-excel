import openpyxl
import string


class Workbook:

    def __init__(self, nombreDePlanilla):
        self.planilla = openpyxl.open(nombreDePlanilla)
        self.nombreDePlanilla = nombreDePlanilla
        self.hojasDePLanilla = {}
        self.armarHojas()
        self.columnas = {
            'Ciudad': 'A',
            'Apellido': 'B',
            'Nombre': 'C',
            'SegundoNombre': 'D',
            'Sexo': 'E',
            'Dni': 'F',
            'Calle': 'G',
            'NumeroDeCalle': 'H',
            'Piso': 'I',
            'Departamento': 'J'
        }

    def armarHojas(self):
        posicionDeHoja = 0
        if 'A' not in self.planilla.sheetnames:
            for i in string.ascii_uppercase:
                self.hojasDePLanilla[i.upper()] = self.planilla.create_sheet(i.upper(), posicionDeHoja)
                posicionDeHoja = posicionDeHoja + 1

    def cargarPersona(self, persona):
        inicialDeCiudad = persona.getCiudad()[0]
        hoja = self.planilla.get_sheet_by_name(inicialDeCiudad.upper())
        fila = hoja.max_row
        for key, value in self.columnas.items():
            hoja['{}{}'.format(value, fila+1)] = eval('persona.get{}()'.format(key))
        self.planilla.save(self.nombreDePlanilla)

    def buscarFila(self, hoja, apellido):
        for row in hoja.rows:
            if row[4].value == "ABC":
                for cell in row:
                    print(cell.value, end=" ")
                print()
    def close(self):
        self.planilla.save(self.nombreDePlanilla)
        self.planilla.close()
